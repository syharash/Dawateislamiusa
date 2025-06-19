import streamlit as st
import openpyxl
from openpyxl import load_workbook
import os

# Safely get path to the Excel file
excel_filename = "Faizan Weekend Girls.xlsx"
excel_path = os.path.join(os.getcwd(), excel_filename)

# Streamlit page title
st.title("📋 Faizan Weekend Girls - Data Entry Form")

# Check if Excel file exists
if not os.path.exists(excel_path):
    st.error(f"Excel file '{excel_filename}' not found in current directory!")
else:
    wb = load_workbook(excel_path)
    sheet = wb["Sheet1"]

    # Determine next row for auto-increment Sl.No
    next_row = sheet.max_row + 1
    sl_no = next_row - 3  # Assuming row 1, 2 and 3 are the headers

    # --- Form Start ---
    with st.form("data_entry_form"):
        st.markdown("### ✏️ Enter New Record")

        campus = st.text_input("Campus", placeholder="e.g. Downtown")
        classes = st.text_input("Classes", placeholder="e.g. KG to 5th")
        employees = st.text_input("Employees", placeholder="e.g. 12")
        students = st.text_input("Current Enrolled Students", placeholder="e.g. 70")
        graduates = st.text_input("Graduates", placeholder="e.g. 15")

        submitted = st.form_submit_button("➕ Submit Record")

        if submitted:
            if not campus or not classes:
                st.warning("Please complete all required fields: Campus and Classes.")
            else:
                try:
                    emp_val = int(employees) if employees else 0
                    stu_val = int(students) if students else 0
                    grad_val = int(graduates) if graduates else 0

                    sheet.cell(row=next_row, column=1, value=sl_no)
                    sheet.cell(row=next_row, column=2, value=campus)
                    sheet.cell(row=next_row, column=3, value=classes)
                    sheet.cell(row=next_row, column=4, value=emp_val)
                    sheet.cell(row=next_row, column=5, value=stu_val)
                    sheet.cell(row=next_row, column=6, value=grad_val)

                    wb.save(excel_path)
                    st.success(f"✅ Record saved at row {next_row}!")
                except ValueError:
                    st.error("Please ensure numeric fields contain valid numbers.")